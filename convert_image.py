import argparse
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
from PIL import Image

def main(args):
    image_filepath:str=args.image_filepath
    output_filepath:str=args.output_filepath
    row_height:float=args.row_height
    column_width:float=args.column_width
    num_rows:int=args.num_rows
    num_columns:int=args.num_columns

    #Resize image
    im=Image.open(image_filepath)
    im_resized=im.resize((num_columns,num_rows))

    #Create workbook
    wb=openpyxl.Workbook()
    ws:Worksheet=wb.active

    #Set row height and column width
    for i in range(1,num_rows+1):
        ws.row_dimensions[i].height=row_height
    for i in range(1,num_columns+1):
        ws.column_dimensions[ws.cell(row=1,column=i).column_letter].width=column_width

    #Set color of each pixel to cell
    for i in range(1,num_columns+1):
        for j in range(1,num_rows+1):
            pixel=im_resized.getpixel((i-1,j-1))
            ws.cell(row=j,column=i).fill=PatternFill(
                patternType="solid",
                fgColor=f"{pixel[0]:02x}{pixel[1]:02x}{pixel[2]:02x}"
            )

    #Save workbook
    wb.save(output_filepath)

if __name__=="__main__":
    parser=argparse.ArgumentParser()
    parser.add_argument("-i","--image-filepath",type=str)
    parser.add_argument("-o","--output-filepath",type=str,default="output.xlsx")
    parser.add_argument("--row-height",type=float,default=3)
    parser.add_argument("--column-width",type=float,default=0.5)
    parser.add_argument("-nr","--num-rows",type=int,default=256)
    parser.add_argument("-nc","--num-columns",type=int,default=256)
    args=parser.parse_args()

    main(args)
